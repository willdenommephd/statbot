import os
import openpyxl
import logging

from .extract_text_data import extract_text_data
from .extract_table_data import extract_table_data
# Function to update existing Excel file with new year data
def update_excel_with_new_year(driver, file_path, urls_titles_filters, new_year, insert_position='end', middle_filters_as_column=None):
    """
    Update an existing Excel file by adding a new year column.
    The percentage change is recalculated using the earliest and latest year.
    
    Args:
        file_path: Path to existing Excel file
        urls_titles_filters: List of (url, table_title, sheet_title, data_type, geography_method, geography_value, middle_filters, start_year, end_year)
        new_year: The year to add (e.g., '2025')
        insert_position: 'start' to insert at beginning (after column C), 'end' to insert before percentage column (default)
        middle_filters_as_column: List of middle filter names to set as column (otherwise row)
    """
    logging.info(f"Updating Excel file: {file_path} with new year: {new_year}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"\nError: File '{file_path}' does not exist.")
        print("You must run the program in normal mode first to create the file.")
        logging.error(f"File not found: {file_path}")
        return False
    
    # Load existing workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        logging.error(f"Could not load workbook: {e}")
        print(f"\nError loading Excel file: {e}")
        return False
    
    # Track which sheets were updated
    updated_sheets = []
    missing_sheets = []
    
    try:
        for entry in urls_titles_filters:
            url, table_title, sheet_title, data_type, geography_method, geography_value, middle_filters, start_year, end_year = entry
            
            # Check if sheet exists
            if sheet_title[:31] not in workbook.sheetnames:
                logging.warning(f"Sheet '{sheet_title}' not found in workbook. Skipping.")
                missing_sheets.append(sheet_title)
                continue
            
            logging.info(f"Processing sheet: {sheet_title}")
            sheet = workbook[sheet_title[:31]]
            
            # Extract new year data with filters set to extract only the new year
            logging.info(f"Extracting data for year {new_year} from {url}")
            if data_type == 'text':
                new_data = extract_text_data(driver, url, geography_method, geography_value, middle_filters, new_year, new_year, middle_filters_as_column=middle_filters_as_column)
            else:
                new_data = extract_table_data(driver, url, geography_method, geography_value, middle_filters, new_year, new_year, middle_filters_as_column=middle_filters_as_column)
            
            # Find the year header row and percentage change column
            year_header_row = None
            pct_change_col = None
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), 1):
                row_values = [cell.value for cell in row]
                
                # Look for "Percentage change" in the row
                for col_idx, val in enumerate(row_values, 1):
                    if val and isinstance(val, str) and 'Percentage change' in val:
                        year_header_row = row_idx
                        pct_change_col = col_idx
                        logging.info(f"Found percentage change column at row {row_idx}, column {col_idx}")
                        break
                
                if year_header_row:
                    break
            
            if not year_header_row or not pct_change_col:
                logging.warning(f"Could not find percentage change column in sheet '{sheet_title}'. Skipping.")
                continue
            
            # Extract the new year column from new_data
            # Find the year header row in new_data
            new_year_col_data = []
            year_row_in_data = None
            
            for data_row_idx, data_row in enumerate(new_data):
                if any(str(cell) == str(new_year) for cell in data_row if cell):
                    year_row_in_data = data_row_idx
                    # Add the year header
                    new_year_col_data.append(new_year)
                    break
            
            if year_row_in_data is None:
                logging.warning(f"Could not find year {new_year} in extracted data for sheet '{sheet_title}'. Skipping.")
                continue
            
            # Extract data values for the new year from subsequent rows
            for data_row_idx in range(year_row_in_data + 1, len(new_data)):
                data_row = new_data[data_row_idx]
                # The year data should be in column index 3 (column D) based on our extraction logic
                if len(data_row) > 3:
                    new_year_col_data.append(data_row[3])  # Column D contains the year data
                else:
                    new_year_col_data.append('')
            
            # Determine where to insert the new year column
            if insert_position == 'start':
                # Insert at the beginning (column 4 = D, right after A, B, C)
                insert_col = 4
                logging.info(f"Inserting new year at the START (column D)")
            else:
                # Insert at the end (before percentage change column)
                insert_col = pct_change_col
                logging.info(f"Inserting new year at the END (before percentage change)")
            
            # Insert the new year column
            sheet.insert_cols(insert_col)
            logging.info(f"Inserted new column at position {insert_col}")
            
            # Write the new year data into the inserted column
            for offset, value in enumerate(new_year_col_data):
                target_row = year_header_row + offset
                if target_row <= sheet.max_row:
                    sheet.cell(row=target_row, column=insert_col, value=value)
            
            # Recalculate percentage change for all data rows
            # After inserting a column, the percentage column shifts right by 1
            if insert_position == 'start':
                # If we inserted at start, pct column moved from pct_change_col to pct_change_col + 1
                new_pct_col = pct_change_col + 1
            else:
                # If we inserted right before pct column, pct column is now at pct_change_col + 1
                new_pct_col = pct_change_col + 1
            
            # Clear the percentage column to recalculate fresh
            for row_idx in range(year_header_row, sheet.max_row + 1):
                sheet.cell(row=row_idx, column=new_pct_col, value='')
            
            # Now recalculate percentage for all rows
            for row_idx in range(year_header_row + 1, sheet.max_row + 1):
                # Skip the header row itself (it's at year_header_row)
                
                # Find first and last numeric values in the row (excluding percentage column)
                first_value = None
                last_value = None
                
                # Data starts at column 4 (D), goes until the new_pct_col (Excel columns are 1-based)
                # Column 1=A, 2=B, 3=C, 4=D (first year), ... new_pct_col-1 (last year), new_pct_col (percentage)
                for col_num in range(4, new_pct_col):
                    cell_value = sheet.cell(row=row_idx, column=col_num).value
                    if cell_value and cell_value != '':
                        try:
                            clean_val = str(cell_value).replace(',', '').replace(' ', '').replace('%', '')
                            num_val = float(clean_val)
                            if first_value is None:
                                first_value = num_val
                            last_value = num_val
                        except (ValueError, AttributeError):
                            pass
                
                # Calculate and write percentage change
                if first_value is not None and last_value is not None and first_value != 0:
                    pct_change = ((last_value - first_value) / first_value) * 100
                    sheet.cell(row=row_idx, column=new_pct_col, value=f"{pct_change:.2f}%")
                else:
                    sheet.cell(row=row_idx, column=new_pct_col, value='')
            
            # Update the header for the percentage column to ensure it's correct
            sheet.cell(row=year_header_row, column=new_pct_col, value='Percentage change from earliest to latest')
            
            updated_sheets.append(sheet_title)
            logging.info(f"Successfully updated sheet: {sheet_title}")
    
    except KeyboardInterrupt:
        print("\n\nUpdate cancelled by user. Saving partial results...")
        logging.info("Update cancelled by user")
    
    # Save the updated workbook
    try:
        workbook.save(file_path)
        print(f"\n✓ Updated {len(updated_sheets)} sheet(s): {', '.join(updated_sheets)}")
        if missing_sheets:
            print(f"\n⚠ Skipped {len(missing_sheets)} missing sheet(s): {', '.join(missing_sheets)}")
            print("  These sheets don't exist in the file. Run in normal mode first.")
        print(f"\nResults saved to: {file_path}")
        return True
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")
        print(f"\nError saving file: {e}")
        return False