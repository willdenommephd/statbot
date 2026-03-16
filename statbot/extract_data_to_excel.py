import os
import openpyxl
import logging

from .add_data_to_sheet import add_data_to_sheet
from .extract_text_data import extract_text_data
from .extract_table_data import extract_table_data
from .open_file import open_file
from .apply_filters_exact import apply_filters_exact

# Function to extract data from multiple links and save to an Excel file
def extract_data_to_excel(driver, urls_titles_filters, file_path, middle_filters_as_column=None):
    logging.info(f"Extracting data to Excel file: {file_path}")
    # Load existing workbook if file already exists (for multi-keyword combined output), otherwise create new
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        logging.info(f"Appending to existing workbook: {file_path}")
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

    try:
        for entry in urls_titles_filters:
            url, table_title, sheet_title, data_type, geography_method, geography_value, middle_filters, start_year, end_year = entry
            logging.debug(f"Unpacked entry - URL: '{url}', sheet: '{sheet_title}'")
            logging.info(f"Processing {sheet_title} ({url})")
            if data_type == 'text':
                data = extract_text_data(driver, url, geography_method, geography_value, middle_filters, start_year, end_year, middle_filters_as_column=middle_filters_as_column)
            else:
                data = extract_table_data(driver, url, geography_method, geography_value, middle_filters, start_year, end_year, middle_filters_as_column=middle_filters_as_column)
            # Deduplicate sheet name if it already exists in a combined workbook
            base_title = sheet_title[:31]
            unique_title = base_title
            counter = 1
            while unique_title in workbook.sheetnames:
                suffix = f"_{counter}"
                unique_title = base_title[:31 - len(suffix)] + suffix
                counter += 1
            sheet = workbook.create_sheet(title=unique_title)
            add_data_to_sheet(sheet, data, table_title, middle_filters_as_column=middle_filters_as_column)
    except KeyboardInterrupt:
        print("\n\nProcess cancelled by user. Saving partial results...")
        logging.info("Process cancelled by user")

    # Ensure at least one sheet is visible
    if not workbook.sheetnames:
        workbook.create_sheet(title="Default Sheet")

    # Save the workbook
    workbook.save(file_path)
    print(f"\nResults saved to: {file_path}")

# Function to extract data from multiple links and save to an Excel file in batches
def extract_data_to_excel_in_batches(driver, urls_titles_filters, file_path, batch_size, middle_filters_as_column=None):
    logging.info(f"Extracting data to Excel file in batches: {file_path}")
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    try:
        for i in range(0, len(urls_titles_filters), batch_size):
            batch = urls_titles_filters[i:i + batch_size]
            for entry in batch:
                url, table_title, sheet_title, data_type, geography_method, geography_value, middle_filters, start_year, end_year = entry
                logging.info(f"Processing {sheet_title} ({url})")
                if data_type == 'text':
                    data = extract_text_data(driver, url, geography_method, geography_value, middle_filters, start_year, end_year, middle_filters_as_column=middle_filters_as_column)
                else:
                    data = extract_table_data(driver, url, geography_method, geography_value, middle_filters, start_year, end_year, middle_filters_as_column=middle_filters_as_column)
                sheet = workbook.create_sheet(title=sheet_title[:31])  # Sheet title max length is 31 characters
                add_data_to_sheet(sheet, data, table_title, middle_filters_as_column=middle_filters_as_column)
            # Optionally save intermediate results
            workbook.save(file_path)
    except KeyboardInterrupt:
        print("\n\nProcess cancelled by user. Saving partial results...")
        logging.info("Process cancelled by user")
    finally:
        # Ensure at least one sheet is visible
        if not workbook.sheetnames:
            workbook.create_sheet(title="Default Sheet")
        print(f"\nResults saved to: {file_path}")
        open_file(file_path)

    # Final save
    workbook.save(file_path)